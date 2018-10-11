import bs4
from bs4 import BeautifulSoup
from Fileconvert import *
import pandas as pd
import numpy as np
from pprint import pprint
import requests
import time
import openpyxl

time0 = time.time()
header = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/65.0.3325.181 Safari/537.36'}


# -------------------------------------------------抓取网页的分割线-------------------------------------------------------
class CKB_html(object):
    def __init__(self, path):
        self.path = path
        self.main_page = 'https://ckb.jax.org'
        self.genes_page = 'https://ckb.jax.org/gene/grid'

    def get_name_url(self, gene):
        """
        输入类型为bs4.element.Tag, 用get()获取嵌套标签中最顶层标签的属性
        :param gene:
        :return:
        """
        url = gene.get('href')
        name = gene.getText().strip()
        return (name, url)

    def get_CKB_genelist(self, url='https://ckb.jax.org/gene/grid'):
        genes_page = requests.post(url, headers=header)
        if genes_page.ok:
            page = genes_page.text
            soup = BeautifulSoup(page, "lxml")
            all_genes = soup.select('a[class="btn btn-default btn-gene btn-block"]')
            return [self.get_name_url(gene) for gene in all_genes]

    # 保存网页到本地
    def save_html(self, url, name):
        """
        对于每个基因页面获取相应的内容，并保存本地
        :param url: 主页链接，str
        :param name: 基因名，str
        :return: None
        """
        web_page = requests.post(url, headers=header)
        if web_page.ok:
            # 获取页面内容，BeautifulSoup解析之，并筛选其中的表格部分
            f = open('{}.html'.format(name), 'w', encoding='utf-8')
            f.write(web_page.text)
            f.close()
            return '{}.html Done!!'.format(name)
        else:
            return None

    # 按基因抓取CKB网页
    def grab_CKB_gene_pages(self, path, foldername, CKB_list):
        """
        根据基因名和链接生成网页链接和保存页面文字，按基因抓取CKB网页
        :param path: 桌面路径
        :param foldername: 保存的文件夹名
        :param CKB_list: 从网页获取的(基因, 基因页面相对路径)
        :return:
        """
        # 创建保存文件夹
        save_path = '{}\\{}'.format(path, foldername)
        try:
            os.mkdir(save_path)
        except FileExistsError:
            pass
        os.chdir(save_path)

        # 记录抓取的基因列表
        CKB_record_file = open('CKB_{}列表.tsv'.format(foldername), 'w', encoding='utf-8')
        CKB_record_file.write('Gene\tURL\tFilename\n')

        # 获取所有基因页面，分别保存
        for each in CKB_list:
            (name, name_url) = each[0:2]
            url = self.main_page + name_url
            if self.save_html(url=url, name=name):
                CKB_record_file.write(tsvline((name, url, '{}.html'.format(name))))
                print('{} Done!!!'.format(name))
            else:
                print('{} Failed!!!'.format(name))

        CKB_record_file.close()


def grab_CKB(path):
    # 创建实例
    ckb = CKB_html(path)
    # 获取需要抓取的基因列表及其链接信息，返回一个由(name，url)构成的列表
    CKB_genelist = ckb.get_CKB_genelist(ckb.genes_page)

    # 开始按照列表中的(name，url)逐一抓取网页保存到本地
    ckb.grab_CKB_gene_pages(path, 'CKB基因网页', CKB_genelist)


# grab_CKB(path)
# -------------------------------------------------解析网页的分割线-------------------------------------------------------


def get_ElementTag(url_file, tag, is_url=True):
    if is_url:
        page = requests.post(url_file, headers=header).text
        # print('Conection OK!!')
    else:
        page = open(url_file, 'r', encoding='utf-8')
    soup = BeautifulSoup(page, 'lxml')
    # 不一定所有的都有tag数据
    tab = soup.select(tag)
    return tab

    # 获取第一个有文本的标签对中的内容,用于获取已经被td或者th标签对圈定的内容中的文本


def get_tag_Text(element_Tag, sep=' '):
    """
    用于获取<td>te<td>text《》<a>65+5+6<th>mmmmmm</th><a></td>xt</td>中的'te', 'text'，td也可以是列表
    td标签对中的a可以被获取但是tr th等不能被获取
    :param element_Tag: 已经被td或者th标签对圈定的文本，可以是列表也可以是bs4.element.Tag
    :return: 均返回一个列表
    """
    if type(element_Tag) == list:
        return [clear_spacer(x.getText(), sep) for x in element_Tag if x]
    elif type(element_Tag) == bs4.element.Tag:
        return [clear_spacer(element_Tag.getText(), sep)]
    else:
        raise TypeError


# 针对行长短不一的表格，选择tag标签对,长度超过给定长度的按给定长度选取, 不足的pass
def get_line_tag(line, tag, length):
    """
    针对line的内容,选择tag标签对,长度超过给定长度的按给定长度选取, 不足的pass
    :param line:line的内容, 类型为 bs4.element.Tag
    :param tag:选择的tag标签对
    :param length:给定长度
    :return: 有则返回文本,无则None
    """
    if len(get_tag_Text(line.select(tag))) >= length:
        return get_tag_Text(line.select(tag))[:length]
    else:
        print(line)
        return None


def append_plus(plus, gene_name, data_type, df_title):
    """
    一行的列表转为df后，第一列插入基因名，后面
    :param plus:输入的需要写入tsv的列
    :param df_title:表头
    :param gene_name:基因名
    :param data_type: 表格归属于variant还是evidence等
    :param add_title: 是否加标题，默认不加
    :return:
    """
    df = pd.DataFrame([gene_name] + plus).T
    df.columns = ['Gene'] + df_title
    if os.path.isfile('.\\CKB_{}.tsv'.format(data_type)):
        add_title = None
    else:
        add_title = True
    df.to_csv('CKB_{}.tsv'.format(data_type), sep='\t', index=None, header=add_title, mode='a+')


# 由<table>标签对圈定的表格，获取表格中的内容,一般性地获取， 标题成为df的columns
def get_table(table_tag):
    """
    获取表格内容函数，包含标题和所有行数据,只采用td的原因在于CKB部分数据是没有tr
    :param table_tag: 由<table>标签对圈定的内容，类型为 bs4.element.Tag
    :return: df
    """
    title = get_tag_Text(table_tag.select('th'))
    title_max = len(title)  # 标题长度
    # 整行直观数据获取，最后插入标题
    content = table_tag.select('td')

    if content == []:
        return 'None'
    else:
        line_max = len(content) // title_max  # 总共内容的行数
        # 整行直观数据获取,td的总数量是title长度的整数倍，故行数为两者之商， 以content[ title_max*i : title_max*(i+1) ]来获取每一行，最后插入标题
        final_table = [get_tag_Text(content[title_max * i: title_max * (i + 1)]) for i in range(line_max)]
        df = pd.DataFrame(final_table)
        df.columns = title
        return df


# 获取标签对内容中的纵向表头的表格
def get_table_T(table_tag, gene_name, data_type, save_path):
    """
    获取表头为首列的表格
    :param table_tag:
    :return:df
    """
    os.chdir(save_path)
    content = table_tag.select('td')
    lines = table_tag.select('tr')
    line_max = len(lines)  # 标题长度
    # 若多了ACMG Incidental List v2.0则不能整除，因为对应的<td>Yes....</td>会重复
    if len(content) % line_max != 0:
        title_max = 2
        print('{} 多了ACMG Incidental List v2.0'.format(gene_name))
    else:
        td_max = len(content)  # 计数所有单元格
        title_max = td_max // line_max  # 总共内容的行数
    # 整行直观数据获取,td的总数量是title长度的整数倍，故行数为两者之商， 以content[ title_max*i : title_max*(i+1) ]来获取每一行，最后插入标题
    final_table = [get_line_tag(line, 'td', title_max) for line in lines if get_line_tag(line, 'td', title_max)]
    # 二维列表转df, 并转置, 转置后， 标题为第一行，内容为第二行，标题替换表头后输出。
    df = pd.DataFrame(final_table).T
    title = list(df.iloc[0, :])
    df = df.iloc[1:, :]
    df.columns = title
    if 'ACMG Incidental List v2.0:' not in df.columns:
        df.insert(3, 'ACMG Incidental List v2.0:', 'NaN')

    if os.path.isfile('.\\CKB_{}.tsv'.format(data_type)):
        add_title = None
    else:
        add_title = True
    df.to_csv('CKB_{}.tsv'.format(data_type), sep='\t', index=None, header=add_title, mode='a+')
    return 1


# 搜索variant网页上的表格，获取转录本坐标信息
def select_loci_byNM(url, gene_name, Refseq_dict, Variant_dict, existed_gene_dict):
    """
    搜索网页上的表格，获取转录本坐标信息
    :param url: variant的网址
    :param gene_name: 基因名
    :param Refseq_dict: ｛基因名：NM｝字典
    :return: 返回一行变异坐标的list
    """
    var_page = requests.post(url, headers=header)
    # print(gene_name, url, 'Variant Conection OK!!!')
    var_soup = BeautifulSoup(var_page.text, 'lxml')
    # 不一定所有的都有转录本数据
    transcript_tab = var_soup.select(
        'table[class="table table-bordered table-hover table-striped basicDataTable transcript-tab-table"]')
    loci = [''] * 6
    if transcript_tab:
        # 表头为纵列的变异介绍表格的第一行第二列即是ckb_NM，故获取该表格所有td选第二个
        ckb_NM = existed_gene_dict[gene_name]
        # 筛选转录本， cura精选转录本与ckb一致则无妨，不一致显示
        try:
            if Refseq_dict[gene_name] != ckb_NM:
                print('CKB_NM: ', ckb_NM, Refseq_dict[gene_name], gene_name)
        except KeyError:
            print('本地这个基因{}没有转录本记录'.format(gene_name), ckb_NM)

        # 根据转录本筛选坐标，若精选转录本，则选精选转录本行，若无，则选CKB转录本行
        NM = get_table(transcript_tab[0])
        title_NM_dict = {y: x for x, y in enumerate(NM.columns)}
        lines_NM = NM.values.tolist()
        lines_NM_dict = {x[0]: x for x in lines_NM}
        # 可能存在NM不属于以上两者,
        try:
            loci = lines_NM_dict[ckb_NM]
        except KeyError:
            loci = lines_NM[0]
        # cDNA和AA去掉c.和p.
        cDNA_num, Protein_num = title_NM_dict['cDNA'], title_NM_dict['Protein']
        loci[cDNA_num] = loci[cDNA_num].split('.')[1]
        loci[Protein_num] = loci[Protein_num].split('.')[1]
        # 更新字典
        Variant_dict[(gene_name, loci[Protein_num])] = loci
        return loci
    else:
        return loci


# 通过molecular_profile的网页链接 获取具体的组合
def get_mol_pro_list(profile, Mol_Pro_dict, main_page):
    """
    输入<td>标签对圈定的profile内容，类型为bs4.element.Tag。
    :param profile:
    :return: mol_pro的str
    """
    profile_text = profile.text.strip()
    try:
        mol_pro_text = Mol_Pro_dict[profile_text]
    except KeyError:
        if profile_text.count(' ') > 1:
            url = main_page + profile.a.get('href')
            deltail_tab = get_ElementTag(url, 'table[class="table table-striped table-hover"]')
            all_tds = deltail_tab[0].select('td')[3]  # 所有的pro都在<a>标签对里
            all_mol_pro = all_tds.select('a')
            mol_pro_list = [a.text.strip().split(' (')[0]
                            for a in all_mol_pro]  # 获取所有的突变, 去掉括号里的内容
            mol_pro_text = ' | '.join(mol_pro_list)
        else:
            mol_pro_text = profile_text
        mol_pro_text.replace(' - ', '-')
        Mol_Pro_dict[profile_text] = mol_pro_text
    return mol_pro_text


def operate_tableline_variant(content, line_max, title_max, main_page, gene_name, data_type, df_title, Refseq_dict,
                              Variant_dict, existed_gene_dict):
    """
    由于表格未必有tr所以采用按标题长度计算td的范围来确定一行，并以此进行操作
    :param content: 整张表格的所有td
    :param line_max: 一共多少行
    :param title_max: 标题有多少列
    :param main_page: CKB主页
    :param gene_name: 基因名
    :param data_type: 表格的数据代表的是variant还是evidence还是profile
    :param Refseq_dict: NM转录本字典，来自于本地文件
    :param Variant_dict: {(gene, aa):[loci]}
    :param df_title: 根据datatype来的title
    :return:
    """
    for i in range(line_max):
        start, stop = title_max * i, title_max * (i + 1)
        line = content[start:stop]

        variant = line[0]
        variant_text = variant.text.strip().replace(' - ', '-')
        query_var_name = (gene_name, variant_text)
        try:
            loci = Variant_dict[query_var_name]
            continue
        except KeyError:
            url = main_page + variant.a.get('href')

            if (variant_text.isupper() and '-' not in variant_text) or ('del' in variant_text
                                                                        or 'ins' in variant_text or 'dup' in variant_text) or 'fs' in variant_text:
                loci = select_loci_byNM(url=url, gene_name=gene_name, Refseq_dict=Refseq_dict,
                                        Variant_dict=Variant_dict, existed_gene_dict=existed_gene_dict)
            else:
                loci = [''] * 6
            plus = get_tag_Text(line) + loci
            plus[1] = variant_text
            append_plus(plus, gene_name=gene_name, data_type=data_type, df_title=df_title)
    return 1


def operate_tableline_profile(content, line_max, title_max, main_page, gene_name, data_type, df_title, Mol_Pro_dict,
                              existed_profile_dict):
    """
    输入<table>标签对圈定的表格内容，类型为bs4.element.Tag。返回一张二维列表
    :param content: 表格， bs4.element.Tag
    :param line_max: 表格共多少行，int
    :param title_max: 标题多少列，int
    :param main_page: 主页网址，str
    :param gene_name: 基因名，str
    :param data_type: 数据分类
    :param Mol_Pro_dict: profile字典，dict
    :param local_file_path: 本地文件路径
    :param final_table: 最终输出表格
    :return:二维列表
    """
    for i in range(line_max):
        start, stop = title_max * i, title_max * (i + 1)
        line = content[start:stop]
        profile = line[0]  # Molecular profile表格中profile的单元格
        get_line = get_tag_Text(line)
        try:
            if existed_profile_dict[tuple([gene_name] + get_line)]:
                continue
        except KeyError:
            plus = get_line + [get_mol_pro_list(profile, Mol_Pro_dict, main_page)]
            append_plus(plus, gene_name=gene_name, data_type=data_type, df_title=df_title)
    return 1


def operate_tableline_evidence(content, line_max, title_max, main_page, gene_name, data_type, df_title, Mol_Pro_dict,
                               Reference_dict, existed_evidence_dict):
    for i in range(line_max):
        start, stop = title_max * i, title_max * (i + 1)
        line = content[start:stop]
        plus = get_tag_Text(line)
        # 获取profile的单元格，从链接替换成profile的组合构成，以‘ | ’隔开
        profile = line[0]
        plus[0] = get_mol_pro_list(profile, Mol_Pro_dict, main_page)
        try:
            if existed_evidence_dict[tuple([gene_name] + plus[:-1])]:
                continue
        except KeyError:
            # 参考文献的具体摘要获取
            refs = line[7].select('a')
            pmids = []
            abstract_union = []
            for ref in refs:
                pmid = ref.text.strip()
                url = '{}{}'.format(main_page, ref.get('href'))
                # 对于有PMID的文献，通过其链接获取摘要；没有的则以url为摘要，以会议的摘要简称作为PMID
                if pmid.isdigit():
                    try:
                        abstract = Reference_dict[pmid]
                    except KeyError:
                        ref_tab = get_ElementTag(url, 'table[class="table table-striped table-hover"]')[0]
                        tds = get_tag_Text(ref_tab.select('td'))
                        abstract = tds[-1]
                        Reference_dict[pmid] = abstract
                    pmids.append(pmid)
                    abstract_union.append(abstract)
                else:
                    ref_tab = get_ElementTag(url, 'table[class="table table-striped table-hover"]')[0]
                    tds = get_tag_Text(ref_tab.select('td'))
                    abstract = tds[-1]
                    title = tds[7]
                    url = tds[-3]
                    if title == 'NCCN.org':
                        pmid = 'NCCN Guideline'
                    elif url != '':
                        Reference_dict[abstract] = url
                        pmid = abstract
                    else:
                        print('这个参考文献没链接：{}'.format(abstract))
                        pmid = abstract
                    pmids.append(pmid)
                    abstract_union.append((url))
            plus[-1] = '|'.join(pmids)
            plus = plus + ['|'.join(abstract_union)]
            append_plus(plus, gene_name=gene_name, data_type=data_type, df_title=df_title)

    return 'ALL {} Done!!'.format(data_type)


def get_table_CKB_simple_way(table_tag, gene_name, data_type, save_path):
    """
    获取表格内容函数，包含标题和所有行数据,只采用td的原因在于CKB部分数据是没有tr
    :param table_tag: 由<table>标签对圈定的内容，类型为 bs4.element.Tag
    :return: df
    """
    print("{} {} Begin...".format(gene_name, data_type))
    # 整行直观数据获取，最后插入标题
    # 整张表格内容不能为空方能继续
    # 整行直观数据获取,td的总数量是title长度的整数倍，故行数为两者之商， 以content[ title_max*i : title_max*(i+1) ]来获取每一行，最后插入标题
    os.chdir(save_path)
    df = get_table(table_tag)
    if type(df) != str:
        df.insert(0, 'Gene', gene_name)
        if os.path.isfile('.\\CKB_{}.tsv'.format(data_type)):
            add_title = None
            mode = 'w'
        else:
            add_title = True
            mode = 'a+'
        df.to_csv('CKB_{}.tsv'.format(data_type), sep='\t', index=None, header=add_title, mode=mode)
        print('{} {} Done!!'.format(gene_name, data_type))


# 获取表格内容函数，包含标题和所有行数据,只采用td的原因在于CKB部分数据是没有tr
def get_table_CKB(table_tag, gene_name, data_type, save_path, Refseq_dict, Variant_dict, Mol_Pro_dict, Reference_dict,
                  existed_gene_dict,
                  existed_profile_dict, existed_evidence_dict):
    """
    获取表格内容函数，包含标题和所有行数据,只采用td的原因在于CKB部分数据是没有tr
    :param table_tag: 由<table>标签对圈定的内容，类型为 bs4.element.Tag
    :return: df
    """
    print("{} {} Begin...".format(gene_name, data_type))
    title = get_tag_Text(table_tag.select('th'))
    title_max = len(title)  # 标题长度
    main_page = 'https://ckb.jax.org'
    # 整行直观数据获取，最后插入标题
    content = table_tag.select('td')
    if content != []:
        # 整张表格内容不能为空方能继续
        line_max = len(content) // title_max  # 总共内容的行数

        # 整行直观数据获取,td的总数量是title长度的整数倍，故行数为两者之商， 以content[ title_max*i : title_max*(i+1) ]来获取每一行，最后插入标题
        os.chdir(save_path)
        if data_type == 'variant':
            df_title = title + ['Transcript', 'gDNA', 'cDNA', 'Protein', 'Source Database',
                                'Genome Build']
            operate_tableline_variant(content=content, line_max=line_max, title_max=title_max, gene_name=gene_name,
                                      main_page=main_page, data_type=data_type, df_title=df_title,
                                      Refseq_dict=Refseq_dict, Variant_dict=Variant_dict,
                                      existed_gene_dict=existed_gene_dict)
        elif data_type == 'profile':
            df_title = title + ['Profile']
            operate_tableline_profile(content=content, line_max=line_max, title_max=title_max, gene_name=gene_name,
                                      main_page=main_page, data_type=data_type, df_title=df_title,
                                      Mol_Pro_dict=Mol_Pro_dict, existed_profile_dict=existed_profile_dict)
        elif data_type == 'evidence':
            df_title = title + ['Abstract']
            operate_tableline_evidence(content=content, line_max=line_max, title_max=title_max, gene_name=gene_name,
                                       main_page=main_page, data_type=data_type, df_title=df_title,
                                       Mol_Pro_dict=Mol_Pro_dict, Reference_dict=Reference_dict,
                                       existed_evidence_dict=existed_evidence_dict)
        print('{} {} Done!!'.format(gene_name, data_type))


def get_table_CKB_recursive(table_tag, gene_name, data_type, save_path, Refseq_dict, Variant_dict, Mol_Pro_dict,
                            Reference_dict, existed_gene_dict, existed_profile_dict, existed_evidence_dict):
    try:
        get_table_CKB(table_tag, gene_name, data_type, save_path, Refseq_dict, Variant_dict, Mol_Pro_dict,
                      Reference_dict, existed_gene_dict, existed_profile_dict, existed_evidence_dict)
    except:
        time.sleep(20)
        return get_table_CKB_recursive(table_tag, gene_name, data_type, save_path, Refseq_dict, Variant_dict,
                                       Mol_Pro_dict, Reference_dict, existed_gene_dict, existed_profile_dict,
                                       existed_evidence_dict)


def getCKB_content(save_path, gene_name, address, Refseq_dict, Variant_dict, Mol_Pro_dict, Reference_dict,
                   existed_gene_dict, existed_profile_dict, existed_evidence_dict, table_rank_dict, local_file_path,
                   is_local_file, simple_way):
    """
    对于每个基因页面获取相应的内容
    :param file_path: 文件路径
    :param save_path: 保存路径
    :param name: 基因名，str
    :param address: 链接，str; 文件名，str 二者均可
    :param is_local_file: 本地与否
    :return: 提示性字符串
    """
    # 是否有本地文件
    if is_local_file:
        os.chdir(local_file_path)
        text = open('{}.html'.format(gene_name), 'r', encoding='utf-8')
    else:
        each_gene_page = requests.post(address, headers=header)
        if each_gene_page.ok:
            text = each_gene_page.text
        else:
            print('Connect {}--{} Failed !!!'.format(gene_name, address))
            return None

    # 切换到保存路径，获取页面内容，BeautifulSoup解析之，并筛选其中的表格部分
    # 获取所有表格后，按固定顺序排序，0,1,2,3 分别代表Gene,Variant,Molecular Profile和Evidence,
    # 考虑到molecular profile的信息可以用到evidence里,profile的赋值优先于evidence

    soup = BeautifulSoup(text, "lxml")
    tables = soup.select('table[class]')
    tables.sort(key=lambda x: table_rank_dict[x.get('class')[-1]][0], reverse=False)

    # 根据data_type进行相应表格处理
    for table in tables:
        # 获取table标签的class属性, 据此判断数据类型
        data_type = table_rank_dict[table.get('class')[-1]][1]
        if data_type == 'gene':
            try:
                if existed_gene_dict[gene_name]:
                    pass
            except KeyError:
                get_table_T(table_tag=table, gene_name=gene_name, data_type=data_type, save_path=save_path)
        elif simple_way == 1:
            get_table_CKB_simple_way(table_tag=table, gene_name=gene_name, data_type=data_type, save_path=save_path)
        else:
            get_table_CKB_recursive(table_tag=table, gene_name=gene_name, data_type=data_type, save_path=save_path,
                                    Refseq_dict=Refseq_dict, Variant_dict=Variant_dict, Mol_Pro_dict=Mol_Pro_dict,
                                    Reference_dict=Reference_dict, existed_gene_dict=existed_gene_dict,
                                    existed_profile_dict=existed_profile_dict,
                                    existed_evidence_dict=existed_evidence_dict)
    return 'All Done!!!'


path = 'C://Users/Administrator/Desktop'
des_path = path


# 从evidence中提取｛PMID: abstract｝字典
def get_ref_dict(file='CKB_evidence.tsv', path='C://Users/Administrator/Desktop'):
    fc = Fileconvert(path)
    ref_dict = {}
    for x in fc.tsv2list(file):
        pmid = x[8]
        abstr = x[9]
        if pmid.isdigit():
            ref_dict[pmid] = abstr

        # 针对多篇文献叠加在一起的
        elif '|' in pmid:
            pmids = pmid.split('|')
            abstrs = abstr.split('|')
            num = 0
            for x in pmids:
                try:
                    if ref_dict[x]:
                        pass
                except KeyError:
                    if x.isdigit():
                        ref_dict[x] = abstrs[num]
                num += 1
            ref_dict[pmid] = abstr
    return ref_dict


def main(simple_way=True):
    # 转录本字典, Molecular_Profile字典, 变异hg38坐标字典
    fc = Fileconvert(des_path)
    Refseq_dict = {x[0]: x[-1] for x in fc.tsv2list('基因介绍列表-merged-merged.tsv')}
    try:
        if simple_way:
            Mol_Pro_dict = {}
            Variant_dict = {}
            Reference_dict = {}
            existed_gene_dict = {}
            existed_profile_dict = {}
            existed_evidence_dict = {}
        else:
            Mol_Pro_dict = {x[1]: x[-1] for x in fc.tsv2list('CKB_profile.tsv')}
            Variant_dict = {(x[0], x[1]): x[5:] for x in fc.tsv2list('CKB_variant.tsv') if set(x[5:]) != {''}}
            Reference_dict = get_ref_dict(file='CKB_evidence.tsv', path=des_path)
            existed_gene_dict = {x[0]: x[-1] for x in fc.tsv2list('CKB_gene.tsv')}
            existed_profile_dict = {tuple(x[:-1]): 1 for x in fc.tsv2list('CKB_profile.tsv')}
            existed_evidence_dict = {tuple(x[:-2]): 1 for x in fc.tsv2list('CKB_evidence.tsv')}
    except FileNotFoundError:
        Mol_Pro_dict = {}
        Variant_dict = {}
        Reference_dict = {}
        existed_gene_dict = {}
        existed_profile_dict = {}
        existed_evidence_dict = {}
    table_rank_dict = {'table-hover': (0, 'gene'),
                       'gene_variant_tab_table': (1, 'variant'),
                       'molecular-profile-tab-table': (2, 'profile'),
                       'profile-response-table-without-treatment-approach': (3, 'evidence'), }

    # 创建保存文件夹
    foldername = 'CKB基因网页'
    local_file_path = '{}\\{}'.format(des_path, foldername)
    try:
        os.mkdir(local_file_path)
    except FileExistsError:
        pass

    done_list = ['ABL1', 'AKT1', 'ALK', 'APC', 'ASXL1', 'ATM', 'ATRX', 'BCOR', 'BCORL1', 'BRAF', 'BRCA1', 'BRCA2',
                 'CALR', 'CBL', 'CBLB', 'CBLC', 'CDH1', 'CDKN2A', 'CEBPA', 'CSF1R', 'CSF3R', 'CTNNB1', 'DNMT3A', 'EGFR',
                 'EML4', 'ERBB2', 'ERBB4', 'ETV6', 'EZH2', 'FBXW7', 'FGFR1', 'FGFR2', 'FGFR3', 'FLT3', 'FOXL2', 'GATA1',
                 'GATA2', 'GNA11', 'GNAQ', 'GNAS', 'HNF1A', 'HRAS', 'IDH1', 'IDH2', 'IKZF1', 'JAK2', 'JAK3', 'KDM6A',
                 'KDR', 'KIT', 'KMT2A', 'KRAS', 'MAP2K1', 'MET', 'MLH1', 'MPL', 'MSH6', 'MYD88', 'NOTCH1', 'NPM1',
                 'NRAS', 'PDGFRA', 'PHF6', 'PIK3CA', 'PTEN', 'PTPN11', 'RAD21', 'RB1', 'RET', 'ROS1', 'RUNX1', 'SETBP1',
                 'SF3B1', 'SMAD4', 'SMARCB1', 'SMC3', 'SMO', 'SRC', 'SRSF2', 'STAG2', 'STK11', 'TET2', 'TP53']
    # 所有CKB基因的网页文件都保存在一个\CKB基因网页的文件夹里, 根据需要创建子目录
    fc = Fileconvert('{}\\CKB基因网页'.format(path))
    # 获取CKB的{基因: 基因html文件名称}字典
    CKB_genelist = [(x[0], x[2]) for x in fc.tsv2list('{}.tsv'.format(foldername))]
    for each in CKB_genelist:
        (gene_name, gene_html) = each
        # if gene_name in done_list:   continue
        getCKB_content(save_path=des_path, gene_name=gene_name, address='{}.html'.format(gene_name),
                       Refseq_dict=Refseq_dict, Variant_dict=Variant_dict, Mol_Pro_dict=Mol_Pro_dict,
                       Reference_dict=Reference_dict, existed_gene_dict=existed_gene_dict,
                       existed_profile_dict=existed_profile_dict, existed_evidence_dict=existed_evidence_dict,
                       table_rank_dict=table_rank_dict, is_local_file=True, local_file_path=local_file_path,
                       simple_way=simple_way)

        print('{} Done!!!'.format(gene_name))
        # 本地文件简易处理不需要间歇10s
        if simple_way:
            pass
        else:
            time.sleep(10)


main(simple_way=True)


# 处理 PMID后置问题
def ckb_translation(ckbfile):
    wb_source = openpyxl.load_workbook(ckbfile)
    sheet_Translation = wb_source['Sheet1']
    for i in range(2, sheet_Translation.max_row + 1):
        chinese = sheet_Translation.cell(row=i, column=2).value.strip().split('|')
        new_text = []
        for y in chinese:
            search = re.search(r'(.+)(（PMID：\d+）)(.+)', y)
            if search:
                text = search.group(1) + search.group(3) + search.group(2)
                new_text.append(text)
            else:
                new_text.append(y)
        sheet_Translation.cell(row=i, column=2).value = '|'.join(new_text)

    wb_source.save('CKB证据翻译.xlsx')


# ckb_translation('ckb证据翻译 (1).xlsx')

time1 = time.time()
print("运行总共耗时为%a秒" % (time1 - time0))
